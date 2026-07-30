#!/usr/bin/env python3
"""Research-group dashboard generator.

Reads a directory of per-person "Research Record" spreadsheets (copies of
``template/FirstnameLastname_Info_Template.xlsx``) and renders a single,
self-contained, interactive static HTML dashboard.

Usage
-----
    python generate_dashboard.py <input_dir> [-o dashboard.html] [--title "..."]

Each ``.xlsx`` in ``<input_dir>`` is one group member. The output is a single
HTML file with all data embedded — no server, no external assets, works by
double-clicking. Interactivity (theme toggle, theme filters, sortable tables)
is plain vanilla JavaScript.

Only ``openpyxl`` is required (see requirements.txt).
"""

from __future__ import annotations

import argparse
import datetime as _dt
import glob
import json
import os
import re
import sys
from typing import Any

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required. Install it with:  pip install openpyxl")


# --------------------------------------------------------------------------- #
#  Spreadsheet parsing
# --------------------------------------------------------------------------- #

TERM_FRACTION = {
    "winter": 0.05,
    "spring": 0.12,
    "summer": 0.45,
    "fall": 0.70,
    "autumn": 0.70,
}


def _clean(value: Any) -> str:
    """Normalise a cell value to a trimmed string ('' for blanks)."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def to_decimal_year(value: Any) -> float | None:
    """Best-effort conversion of a term/date string to a decimal year.

    Handles datetimes, ``2024``, ``2024-03``, ``2024 Fall`` / ``Fall 2024`` and
    a few similar shapes. Returns ``None`` when no year can be found.
    """
    if value is None:
        return None
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.year + (value.month - 1) / 12.0

    s = _clean(value)
    if not s:
        return None

    # find a 4-digit year anywhere in the string
    m = re.search(r"((?:19|20)\d{2})", s)
    if not m:
        return None
    year = int(m.group(1))

    # explicit month number after the year (2024-03)
    m2 = re.search(r"(?:19|20)\d{2}[-/](\d{1,2})", s)
    frac = None
    if m2:
        month = int(m2.group(1))
        if 1 <= month <= 12:
            frac = (month - 1) / 12.0

    if frac is None:
        low = s.lower()
        for term, f in TERM_FRACTION.items():
            if term in low:
                frac = f
                break

    return year + (frac if frac is not None else 0.5)


def find_label_row(ws, text: str, col: str = "A") -> int | None:
    """Return the 1-based row where column ``col`` equals ``text`` (case-insensitive)."""
    target = text.strip().lower()
    for r in range(1, ws.max_row + 1):
        v = ws[f"{col}{r}"].value
        if v is not None and str(v).strip().lower() == target:
            return r
    return None


def _row_values(ws, row: int, ncols: int) -> list[str]:
    return [_clean(ws.cell(row=row, column=c).value) for c in range(1, ncols + 1)]


def read_table(ws, header_row: int, ncols: int, stop_labels=()) -> list[list[str]]:
    """Read data rows below ``header_row``.

    Skips the example row (column A == 'ex'), fully blank rows, and stops at any
    row whose first column matches one of ``stop_labels``.
    """
    stops = {s.lower() for s in stop_labels}
    out: list[list[str]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        first = _clean(ws.cell(row=r, column=1).value).lower()
        if first in stops:
            break
        vals = _row_values(ws, r, ncols)
        if first == "ex":
            continue
        # ignore the leading '#' column when testing for emptiness
        if all(v == "" for v in vals[1:]):
            continue
        out.append(vals)
    return out


def parse_projects(raw: str) -> list[dict]:
    """Split a 'Primary project(s)' cell into individual projects.

    Splits on newlines / semicolons / commas and flags '(lead)'.
    """
    projects = []
    if not raw:
        return projects
    parts = re.split(r"[\n;,]+", raw)
    for part in parts:
        name = part.strip()
        if not name:
            continue
        is_lead = False
        m = re.search(r"\(\s*lead\s*\)", name, re.I)
        if m:
            is_lead = True
            name = (name[: m.start()] + name[m.end():]).strip()
        name = name.strip(" -–·")
        if name:
            projects.append({"name": name, "lead": is_lead})
    return projects


def parse_workbook(path: str) -> dict:
    """Parse one Research Record workbook into a structured person dict."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    names = {n.lower(): n for n in wb.sheetnames}

    def sheet(*candidates):
        for c in candidates:
            if c.lower() in names:
                return wb[names[c.lower()]]
        return None

    person: dict[str, Any] = {"file": os.path.basename(path)}

    # ---- Overview ----------------------------------------------------------
    ov = sheet("Overview")
    if ov is not None:
        def val(label):
            r = find_label_row(ov, label)
            return _clean(ov[f"B{r}"].value) if r else ""

        person["rank"] = val("Rank")
        person["name"] = val("Full name")
        person["theme"] = val("Research theme / area")
        person["program_start"] = val("Program start (year / term)")
        person["group_start"] = val("Started with the group")
        person["program_year"] = val("Current program year")
        person["grad_term"] = val("Expected graduation term")
        person["effort"] = val("Appointment / effort")
        person["site"] = val("Primary site / base")
        person["funding"] = val("Primary funding source")
        person["thesis_title"] = val("Working thesis title / topic")
        person["projects_raw"] = val("Primary project(s)")
        person["focus"] = val("Current focus / status")

        # milestones (Status / Date / Location / Notes in cols B/C/D/E)
        milestones = {}
        for key, label in [
            ("qual", "Qualifying exam"),
            ("comp", "Comprehensive exam"),
            ("committee", "Thesis committee formed"),
            ("author", "Author qualification (if applicable)"),
            ("defense", "Thesis defense"),
        ]:
            r = find_label_row(ov, label)
            if r:
                milestones[key] = {
                    "status": _clean(ov[f"B{r}"].value),
                    "date": _clean(ov[f"C{r}"].value),
                    "loc": _clean(ov[f"D{r}"].value),
                    "notes": _clean(ov[f"E{r}"].value),
                }
            else:
                milestones[key] = {"status": "", "date": "", "loc": "", "notes": ""}
        person["milestones"] = milestones

    person.setdefault("name", "")
    if not person.get("name"):
        # fall back to the filename (strip _Info_Template etc.)
        base = os.path.splitext(person["file"])[0]
        base = re.sub(r"[_\- ]*info.*$", "", base, flags=re.I)
        person["name"] = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", base) or base

    person["projects"] = parse_projects(person.get("projects_raw", ""))
    person["start_year"] = to_decimal_year(person.get("program_start"))
    person["group_year"] = to_decimal_year(person.get("group_start")) or person["start_year"]

    # ---- Committee ---------------------------------------------------------
    com = sheet("Committee")
    person["committee_members"] = []
    person["committee_meetings"] = []
    person["progress_reports"] = []
    if com is not None:
        r_mem = find_label_row(com, "Committee members")
        r_meet = find_label_row(com, "Committee meetings")
        r_prog = find_label_row(com, "Annual progress reports")
        if r_mem:
            for v in read_table(com, r_mem + 1, 5, stop_labels=("Committee meetings",)):
                person["committee_members"].append(
                    {"name": v[1], "role": v[2], "dept": v[3], "notes": v[4]}
                )
        if r_meet:
            for v in read_table(com, r_meet + 1, 5, stop_labels=("Annual progress reports",)):
                person["committee_meetings"].append(
                    {"meeting": v[0], "date": v[1], "loc": v[2], "focus": v[3], "notes": v[4]}
                )
        if r_prog:
            for v in read_table(com, r_prog + 1, 4):
                person["progress_reports"].append(
                    {"date": v[1], "period": v[2], "notes": v[3]}
                )

    # ---- Roles -------------------------------------------------------------
    person["roles"] = []
    rl = sheet("Roles")
    if rl is not None:
        h = find_label_row(rl, "#")
        if h:
            for v in read_table(rl, h, 7):
                person["roles"].append(
                    {"role": v[1], "category": v[2], "org": v[3],
                     "start": v[4], "end": v[5], "notes": v[6]}
                )

    # ---- Presentations -----------------------------------------------------
    person["presentations"] = []
    pr = sheet("Presentations")
    if pr is not None:
        h = find_label_row(pr, "#")
        if h:
            for v in read_table(pr, h, 9):
                person["presentations"].append(
                    {"type": v[1], "venue": v[2], "location": v[3], "date": v[4],
                     "title": v[5], "invited": v[6], "travel": v[7], "link": v[8]}
                )

    # ---- Outreach ----------------------------------------------------------
    person["outreach"] = []
    ou = sheet("Outreach & education", "Outreach")
    if ou is not None:
        h = find_label_row(ou, "#")
        if h:
            for v in read_table(ou, h, 9):
                person["outreach"].append(
                    {"activity": v[1], "type": v[2], "audience": v[3], "date": v[4],
                     "role": v[5], "hours": v[6], "notes": v[7], "url": v[8]}
                )

    # ---- Publications ------------------------------------------------------
    person["publications"] = []
    pu = sheet("Publications")
    if pu is not None:
        h = find_label_row(pu, "#")
        if h:
            for v in read_table(pu, h, 7):
                person["publications"].append(
                    {"title": v[1], "role": v[2], "journal": v[3], "status": v[4],
                     "date": v[5], "doi": v[6]}
                )

    # ---- Applications ------------------------------------------------------
    person["applications"] = []
    ap = sheet("Applications")
    if ap is not None:
        h = find_label_row(ap, "#")
        if h:
            for v in read_table(ap, h, 8):
                person["applications"].append(
                    {"name": v[1], "type": v[2], "status": v[3], "requested": v[4],
                     "awarded": v[5], "date": v[6], "notes": v[7]}
                )

    # ---- Awards ------------------------------------------------------------
    person["awards"] = []
    aw = sheet("Awards")
    if aw is not None:
        h = find_label_row(aw, "#")
        if h:
            for v in read_table(aw, h, 6):
                person["awards"].append(
                    {"award": v[1], "org": v[2], "date": v[3],
                     "amount": v[4], "notes": v[5]}
                )

    # ---- Conferences & Schools (attendance) --------------------------------
    person["events"] = []
    ev = sheet("Conferences & Schools", "Conferences and Schools", "Conferences")
    if ev is not None:
        h = find_label_row(ev, "#")
        if h:
            for v in read_table(ev, h, 9):
                person["events"].append(
                    {"type": v[1], "event": v[2], "location": v[3], "start": v[4],
                     "end": v[5], "role": v[6], "presented": v[7], "notes": v[8]}
                )

    # ---- Grants & Funding --------------------------------------------------
    person["grants"] = []
    gr = sheet("Grants & Funding", "Grants and Funding", "Grants")
    if gr is not None:
        h = find_label_row(gr, "#")
        if h:
            for v in read_table(gr, h, 9):
                person["grants"].append(
                    {"title": v[1], "agency": v[2], "role": v[3], "amount": v[4],
                     "start": v[5], "end": v[6], "status": v[7], "notes": v[8]}
                )

    wb.close()
    return person


# --------------------------------------------------------------------------- #
#  Aggregation into the dashboard data model
# --------------------------------------------------------------------------- #

# Mid-tone palette that reads on both light and dark backgrounds.
PALETTE = [
    "#9c5b46", "#3f6079", "#5b7a5b", "#9c7b2e",
    "#7a5b8a", "#3f8a8a", "#a5563f", "#6b7a3f",
]


def _theme_key(theme: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", theme.lower()).strip("-") or "other"


def is_postdoc(person: dict) -> bool:
    return "postdoc" in (person.get("rank", "").lower())


def is_faculty(person: dict) -> bool:
    r = person.get("rank", "").lower()
    return ("faculty" in r) or ("professor" in r) or (" pi" in f" {r}") or ("staff" in r)


def is_student(person: dict) -> bool:
    r = person.get("rank", "").lower()
    if is_postdoc(person) or is_faculty(person):
        return False
    return ("grad" in r) or ("student" in r) or ("undergrad" in r)


def parse_money(value) -> float:
    """Pull a number out of a currency-ish string ('$50,000' -> 50000.0)."""
    s = _clean(value)
    if not s:
        return 0.0
    s = re.sub(r"[^0-9.]", "", s)
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def build_model(people: list[dict], title: str) -> dict:
    # ---- themes & colors ---------------------------------------------------
    themes: dict[str, dict] = {}
    order = 0
    for p in people:
        theme = p.get("theme") or "Unspecified"
        key = _theme_key(theme)
        p["key"] = key
        if key not in themes:
            themes[key] = {"key": key, "label": theme, "color": PALETTE[order % len(PALETTE)]}
            order += 1

    # ---- projects & membership --------------------------------------------
    project_key: dict[str, str] = {}
    membership: dict[str, dict] = {}
    for p in people:
        pm = {}
        for proj in p.get("projects", []):
            name = proj["name"]
            project_key.setdefault(name, p["key"])  # colour by first owner's theme
            pm[name] = "lead" if proj["lead"] else "member"
        if pm:
            membership[p["name"]] = pm
    projects = [{"name": n, "key": k} for n, k in project_key.items()]

    # ---- flatten cross-cutting records ------------------------------------
    def flat(field, extra=None):
        rows = []
        for p in people:
            for rec in p.get(field, []):
                item = dict(rec)
                item["who"] = p["name"]
                item["key"] = p["key"]
                rows.append(item)
        return rows

    now = _dt.datetime.now()
    now_dec = now.year + (now.month - 1) / 12.0 + (now.day / 31.0) / 12.0

    n_students = sum(1 for p in people if is_student(p))
    n_postdocs = sum(1 for p in people if is_postdoc(p))
    n_faculty = sum(1 for p in people if is_faculty(p))

    def passed(key):
        c = 0
        for p in people:
            st = p.get("milestones", {}).get(key, {}).get("status", "").lower()
            if st in ("passed", "yes"):
                c += 1
        return c

    apps = flat("applications")
    n_awarded = sum(1 for a in apps if a.get("status", "").lower() == "awarded")

    grants = flat("grants")
    active_grants = sum(1 for g in grants
                        if g.get("status", "").lower() in ("active", "awarded"))
    grant_total = sum(parse_money(g.get("amount"))
                      for g in grants if g.get("status", "").lower() in ("active", "awarded"))

    model = {
        "group": {
            "title": title,
            "generated": now.strftime("%Y-%m-%d"),
        },
        "now": round(now_dec, 3),
        "themes": list(themes.values()),
        "people": people,
        "projects": projects,
        "membership": membership,
        "applications": apps,
        "grants": grants,
        "events": flat("events"),
        "presentations": flat("presentations"),
        "awards": flat("awards"),
        "publications": flat("publications"),
        "roles": flat("roles"),
        "outreach": flat("outreach"),
        "kpis": {
            "people": len(people),
            "students": n_students,
            "postdocs": n_postdocs,
            "faculty": n_faculty,
            "themes": len(themes),
            "sites": len({p.get("site", "") for p in people if p.get("site")}),
            "quals": passed("qual"),
            "comps": passed("comp"),
            "publications": sum(len(p.get("publications", [])) for p in people),
            "presentations": sum(len(p.get("presentations", [])) for p in people),
            "awards": sum(len(p.get("awards", [])) for p in people),
            "awarded": n_awarded,
            "grants": active_grants,
            "grant_total": grant_total,
        },
    }
    return model


# --------------------------------------------------------------------------- #
#  HTML rendering
# --------------------------------------------------------------------------- #

def add_stamp(path: str, stamp: str) -> str:
    """Insert a `_<stamp>` before the extension: dashboard.html -> dashboard_2026-07-30.html."""
    root, ext = os.path.splitext(path)
    return f"{root}_{stamp}{ext}"


def render_html(model: dict) -> str:
    data_json = json.dumps(model, ensure_ascii=False, indent=None)
    # guard against an accidental </script> inside data
    data_json = data_json.replace("</", "<\\/")
    return HTML_TEMPLATE.replace("/*__DATA__*/null", data_json)


def export_pdf(html_path: str, pdf_path: str, chrome: str | None = None) -> bool:
    """Print the rendered dashboard to PDF with headless Chromium (Playwright).

    Playwright runs the dashboard's JavaScript before printing, so the PDF
    matches the live page. Requires:  pip install playwright && playwright install chromium
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  PDF skipped — Playwright not installed.\n"
              "    pip install playwright && playwright install chromium", file=sys.stderr)
        return False
    import pathlib
    url = pathlib.Path(html_path).resolve().as_uri()
    try:
        with sync_playwright() as p:
            launch = {"executable_path": chrome} if chrome else {}
            browser = p.chromium.launch(**launch)
            page = browser.new_page()
            page.goto(url, wait_until="networkidle")
            page.emulate_media(media="print")
            page.pdf(path=pdf_path, format="A4", print_background=True,
                     margin={"top": "0.5in", "bottom": "0.5in",
                             "left": "0.5in", "right": "0.5in"})
            browser.close()
    except Exception as exc:
        print(f"  PDF export failed: {exc}\n"
              "    Try:  playwright install chromium", file=sys.stderr)
        return False
    print(f"Wrote {pdf_path}")
    return True


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("input_dir", nargs="?",
                    help="directory containing the per-person .xlsx records")
    ap.add_argument("-o", "--output", default="dashboard.html", help="output HTML file")
    ap.add_argument("--no-datestamp", action="store_true",
                    help="do not append the generation date to output filenames")
    ap.add_argument("--title", default="Research Group — Dashboard", help="dashboard title")
    ap.add_argument("--pdf", nargs="?", const="__auto__", default=None,
                    metavar="PATH", help="also write a PDF (default: alongside the HTML)")
    ap.add_argument("--chrome", default=os.environ.get("CHROME_BIN"),
                    help="path to a Chrome/Chromium binary for --pdf (else Playwright's own)")
    ap.add_argument("--dump-template", metavar="PATH",
                    help="write the raw HTML template (for the Apps Script port) and exit")
    args = ap.parse_args(argv)

    if args.dump_template:
        with open(args.dump_template, "w", encoding="utf-8") as f:
            f.write(HTML_TEMPLATE)
        print(f"Wrote template to {args.dump_template}")
        return 0

    if not args.input_dir:
        ap.error("input_dir is required (unless using --dump-template)")
    if not os.path.isdir(args.input_dir):
        sys.exit(f"Not a directory: {args.input_dir}")

    paths = sorted(
        p for p in glob.glob(os.path.join(args.input_dir, "*.xlsx"))
        if not os.path.basename(p).startswith(("~$", "."))
    )
    if not paths:
        sys.exit(f"No .xlsx files found in {args.input_dir}")

    people = []
    for path in paths:
        try:
            people.append(parse_workbook(path))
            print(f"  parsed  {os.path.basename(path)}")
        except Exception as exc:  # keep going; one bad file shouldn't kill the run
            print(f"  SKIP    {os.path.basename(path)}: {exc}", file=sys.stderr)

    if not people:
        sys.exit("No records could be parsed.")

    model = build_model(people, args.title)
    html = render_html(model)

    stamp = model["group"]["generated"]  # YYYY-MM-DD, same date shown in the dashboard
    out_html = args.output if args.no_datestamp else add_stamp(args.output, stamp)
    with open(out_html, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"\nWrote {out_html}  ({len(people)} member record(s), "
          f"{len(model['themes'])} theme(s))")

    if args.pdf is not None:
        if args.pdf != "__auto__":
            pdf_path = args.pdf if args.no_datestamp else add_stamp(args.pdf, stamp)
        else:
            pdf_path = os.path.splitext(out_html)[0] + ".pdf"
        export_pdf(out_html, pdf_path, chrome=args.chrome)
    return 0


# --------------------------------------------------------------------------- #
#  The HTML/CSS/JS template  (data is injected in place of /*__DATA__*/null)
# --------------------------------------------------------------------------- #

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Research Group — Dashboard</title>
<style>
  :root{
    --paper:#fbf8f1; --ink:#22201c; --ink-soft:#6b665d; --rule:#d9d3c6;
    --panel:#fffdf8; --band:rgba(150,120,60,.10); --present:#b0483b;
    --secured:#3f6079; --applied:#9c7b2e; --planned:#8a8577; --pending:#a5563f;
    --good:#4d7a52; --warn:#b08a2e; --open:#a5563f;
    --font-serif:"Iowan Old Style","Palatino Linotype",Palatino,Georgia,serif;
    --font-sans:-apple-system,BlinkMacSystemFont,"Segoe UI",Helvetica,Arial,sans-serif;
  }
  html[data-theme="dark"]{
    --paper:#15171c; --ink:#e6e3db; --ink-soft:#9a968c; --rule:#2f333b;
    --panel:#1b1e24; --band:rgba(200,175,110,.10); --present:#e0725f;
    --secured:#7aa6c8; --applied:#d4ad57; --planned:#9a968c; --pending:#e0725f;
    --good:#7fb583; --warn:#d4ad57; --open:#e0725f;
  }
  *{box-sizing:border-box}
  body{margin:0;background:var(--paper);color:var(--ink);font-family:var(--font-serif);
       line-height:1.5;-webkit-font-smoothing:antialiased;transition:background .3s,color .3s}
  .wrap{max-width:1180px;margin:0 auto;padding:2.4rem 1.6rem 5rem}
  header.top{display:flex;justify-content:space-between;align-items:flex-start;gap:1rem;flex-wrap:wrap;
       border-bottom:1px solid var(--rule);padding-bottom:1rem;margin-bottom:1.6rem}
  h1{font-size:1.9rem;font-weight:600;margin:0 0 .2rem;letter-spacing:-.01em}
  .sub{color:var(--ink-soft);font-family:var(--font-sans);font-size:.85rem;margin:0}
  .controls{display:flex;gap:.5rem;align-items:center;font-family:var(--font-sans)}
  button.ghost{background:none;border:1px solid var(--rule);color:var(--ink-soft);
       padding:.35rem .7rem;border-radius:5px;cursor:pointer;font-size:.8rem}
  button.ghost:hover{border-color:var(--ink-soft);color:var(--ink)}
  .filters{display:flex;flex-wrap:wrap;gap:.4rem;margin-bottom:2rem;font-family:var(--font-sans)}
  .chip{border:1px solid var(--rule);background:none;color:var(--ink-soft);border-radius:20px;
       padding:.3rem .85rem;font-size:.78rem;cursor:pointer;transition:.15s}
  .chip:hover{color:var(--ink)}
  .chip.active{background:var(--ink);color:var(--paper);border-color:var(--ink)}
  .kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(120px,1fr));gap:1px;
       background:var(--rule);border:1px solid var(--rule);border-radius:8px;overflow:hidden;margin-bottom:2.6rem}
  .kpi{background:var(--panel);padding:1rem 1.1rem}
  .kpi .n{font-size:1.9rem;font-weight:600;line-height:1}
  .kpi .l{font-family:var(--font-sans);font-size:.72rem;color:var(--ink-soft);text-transform:uppercase;letter-spacing:.05em;margin-top:.35rem}
  section{margin-bottom:2.8rem}
  h2{font-size:1.15rem;font-weight:600;margin:0 0 .2rem;display:flex;align-items:baseline;gap:.6rem}
  h2 .note{font-family:var(--font-sans);font-size:.72rem;color:var(--ink-soft);font-weight:400}
  .lead{color:var(--ink-soft);font-size:.9rem;margin:.1rem 0 1rem;max-width:64ch}
  .chart{width:100%;overflow-x:auto}
  svg{display:block;max-width:100%}
  .axis text{font-family:var(--font-sans);font-size:10px;fill:var(--ink-soft)}
  .rowlabel{font-family:var(--font-sans);font-size:11.5px;fill:var(--ink)}
  .barcap{font-family:var(--font-sans);font-size:10px;fill:var(--ink-soft)}
  table{border-collapse:collapse;width:100%;font-size:.86rem}
  th{font-family:var(--font-sans);font-weight:600;text-align:left;font-size:.72rem;
     text-transform:uppercase;letter-spacing:.04em;color:var(--ink-soft);
     border-bottom:1.5px solid var(--ink);padding:.5rem .6rem;cursor:pointer;white-space:nowrap}
  th:hover{color:var(--ink)}
  td{padding:.5rem .6rem;border-bottom:1px solid var(--rule);vertical-align:top}
  tr:hover td{background:var(--band)}
  .theme-dot{display:inline-block;width:8px;height:8px;border-radius:50%;margin-right:.45rem;vertical-align:middle}
  .muted{color:var(--ink-soft)}
  .matrix{overflow-x:auto}
  .matrix table{font-size:.82rem}
  .matrix td.name,.matrix th.proj{white-space:nowrap}
  .matrix th.proj{writing-mode:vertical-rl;transform:rotate(180deg);text-align:left;
       height:120px;vertical-align:bottom;padding:.4rem .3rem;border-bottom:1.5px solid var(--ink)}
  .matrix td{text-align:center}
  .matrix td.name{text-align:left;font-family:var(--font-serif);font-size:.9rem}
  .cell{display:inline-block;width:10px;height:10px;border-radius:2px}
  .cell.lead{width:14px;height:14px;border-radius:3px}
  .pill{font-family:var(--font-sans);font-size:.7rem;padding:.15rem .5rem;border-radius:20px;display:inline-block}
  .pill.ok{background:rgba(77,122,82,.14);color:var(--good)}
  .pill.pending{background:rgba(176,138,46,.16);color:var(--warn)}
  .pill.na{color:var(--ink-soft)}
  .board{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:1rem}
  .col{border:1px solid var(--rule);border-radius:8px;background:var(--panel);overflow:hidden}
  .col h3{margin:0;font-family:var(--font-sans);font-size:.74rem;text-transform:uppercase;letter-spacing:.05em;
       padding:.6rem .8rem;border-bottom:1px solid var(--rule);display:flex;justify-content:space-between}
  .col .items{padding:.5rem .55rem;display:flex;flex-direction:column;gap:.4rem}
  .fcard{border-left:3px solid var(--rule);padding:.4rem .55rem;font-size:.84rem}
  .fcard .who{font-weight:600}
  .fcard .what{color:var(--ink-soft);font-family:var(--font-sans);font-size:.76rem}
  .col.awarded h3{color:var(--secured)} .col.awarded .fcard{border-left-color:var(--secured)}
  .col.applied h3{color:var(--applied)} .col.applied .fcard{border-left-color:var(--applied)}
  .col.under-review h3{color:var(--applied)} .col.under-review .fcard{border-left-color:var(--applied)}
  .col.planned h3{color:var(--planned)} .col.planned .fcard{border-left-color:var(--planned)}
  .col.declined h3{color:var(--pending)} .col.declined .fcard{border-left-color:var(--pending)}
  .talks{display:flex;flex-direction:column;gap:.15rem}
  .talk{display:grid;grid-template-columns:80px 1fr auto;gap:.7rem;padding:.45rem 0;border-bottom:1px solid var(--rule);align-items:baseline}
  .talk .yr{font-family:var(--font-sans);font-size:.8rem;color:var(--ink-soft)}
  .talk .prize{color:var(--applied);font-family:var(--font-sans);font-size:.72rem}
  .legend{font-family:var(--font-sans);font-size:.75rem;color:var(--ink-soft);display:flex;gap:1.1rem;flex-wrap:wrap;margin-top:.6rem}
  .legend span{display:inline-flex;align-items:center;gap:.35rem}
  footer{border-top:1px solid var(--rule);margin-top:3rem;padding-top:1rem;
       font-family:var(--font-sans);font-size:.75rem;color:var(--ink-soft)}
  .empty{color:var(--ink-soft);font-family:var(--font-sans);font-size:.82rem;padding:.4rem 0}
  a{color:inherit}

  /* ---------- print / PDF ---------- */
  @page{margin:0.5in}
  @media print{
    :root, html[data-theme="dark"]{
      --paper:#fff; --ink:#22201c; --ink-soft:#5a564d; --rule:#c9c3b6;
      --panel:#fff; --band:rgba(150,120,60,.10);
    }
    body{-webkit-print-color-adjust:exact;print-color-adjust:exact}
    .controls,.filters{display:none!important}
    .wrap{max-width:none;padding:0}
    section{break-inside:avoid;page-break-inside:avoid;margin-bottom:1.6rem}
    h1,h2{break-after:avoid}
    .chart,.matrix,.board,#peopleTable,#milestoneTable,#grantsTable,#pubsTable,
    #rolesTable,#outreachTable,#eventsTable{overflow:visible}
    tr:hover td{background:none}
  }
</style>
</head>
<body>
<div class="wrap">
  <header class="top">
    <div>
      <h1 id="title"></h1>
      <p class="sub" id="subtitle"></p>
    </div>
    <div class="controls"><button class="ghost" id="themeBtn">◐ Theme</button></div>
  </header>

  <div class="filters" id="filters"></div>
  <div class="kpis" id="kpis"></div>

  <section id="sec-cohort">
    <h2>Student cohort timeline <span class="note">program start → present, with completion window</span></h2>
    <p class="lead">Bars run from each student's program start to the present. The shaded band marks the typical 5–6 year PhD completion window.</p>
    <div class="chart" id="cohortChart"></div>
  </section>

  <section id="sec-postdoc">
    <h2>Postdoc tenure <span class="note">years since arrival</span></h2>
    <p class="lead">The band beyond ~3 years flags tenure where a career-stage conversation is due.</p>
    <div class="chart" id="postdocChart"></div>
  </section>

  <section id="sec-people">
    <h2>People <span class="note">click a column header to sort</span></h2>
    <div id="peopleTable"></div>
    <div class="legend" id="legend"></div>
  </section>

  <section id="sec-milestones">
    <h2>Milestones &amp; exams <span class="note">academic status</span></h2>
    <div id="milestoneTable"></div>
  </section>

  <section id="sec-matrix">
    <h2>Who works on what <span class="note">large square = lead</span></h2>
    <div class="matrix" id="matrix"></div>
  </section>

  <section id="sec-grants">
    <h2>Grants &amp; funding <span class="note">the group's funded portfolio</span></h2>
    <div id="grantsTable"></div>
  </section>

  <section id="sec-funding">
    <h2>Applications &amp; fellowship pipeline <span class="note">individual applications, by status</span></h2>
    <div class="board" id="fundingBoard"></div>
  </section>

  <section id="sec-pubs">
    <h2>Publications <span class="note">papers &amp; notes</span></h2>
    <div id="pubsTable"></div>
  </section>

  <section id="sec-talks">
    <h2>Presentations &amp; awards <span class="note">on record</span></h2>
    <div class="talks" id="talks"></div>
  </section>

  <section id="sec-events">
    <h2>Conferences &amp; schools <span class="note">events attended</span></h2>
    <div id="eventsTable"></div>
  </section>

  <section id="sec-roles">
    <h2>Roles &amp; service <span class="note">positions held</span></h2>
    <div id="rolesTable"></div>
  </section>

  <section id="sec-outreach">
    <h2>Outreach &amp; education <span class="note">engagement activities</span></h2>
    <div id="outreachTable"></div>
  </section>

  <footer id="footer"></footer>
</div>

<script>
const DATA = /*__DATA__*/null;

/* ---------- helpers ---------- */
const $=(s,r=document)=>r.querySelector(s);
const el=(t,c,h)=>{const e=document.createElement(t);if(c)e.className=c;if(h!=null)e.innerHTML=h;return e;};
const esc=s=>String(s==null?"":s).replace(/[&<>]/g,m=>({"&":"&amp;","<":"&lt;",">":"&gt;"}[m]));
const NOW=DATA.now;
const THEME=Object.fromEntries(DATA.themes.map(t=>[t.key,t]));
const themeColor=k=>(THEME[k]&&THEME[k].color)||"var(--ink-soft)";
let activeFilter="all";
const match=k=>activeFilter==="all"||activeFilter===k;
const isPostdoc=p=>/postdoc/i.test(p.rank||"");
const isStudent=p=>!isPostdoc(p)&&/(grad|student|undergrad)/i.test(p.rank||"");
const people=DATA.people;
const visiblePeople=()=>people.filter(p=>match(p.key));
const hide=(id,on)=>{const s=$(id);if(s)s.classList.toggle("hidden-sec",on);s.style.display=on?"none":"";};

/* ---------- header ---------- */
$("#title").textContent=DATA.group.title;
$("#subtitle").textContent=`${people.length} member${people.length!==1?"s":""} · `
  +`${DATA.kpis.students} student${DATA.kpis.students!==1?"s":""}, ${DATA.kpis.postdocs} postdoc${DATA.kpis.postdocs!==1?"s":""}`
  +(DATA.kpis.faculty?`, ${DATA.kpis.faculty} faculty/staff`:"")+` · `
  +`${DATA.themes.length} theme${DATA.themes.length!==1?"s":""} · generated ${DATA.group.generated}`;
$("#footer").innerHTML=`Generated from ${people.length} Research Record spreadsheet(s) on ${DATA.group.generated}. `
  +`Each member maintains their own workbook; re-run <code>generate_dashboard.py</code> to refresh.`;

/* ---------- filters ---------- */
function renderFilters(){
  const box=$("#filters");box.innerHTML="";
  const opts=[["all","All"]].concat(DATA.themes.map(t=>[t.key,t.label]));
  opts.forEach(([k,label])=>{
    const b=el("button","chip"+(activeFilter===k?" active":""),esc(label));
    if(k!=="all"){const d=el("span","theme-dot");d.style.background=themeColor(k);b.prepend(d);}
    b.onclick=()=>{activeFilter=k;renderAll();};
    box.append(b);
  });
}

/* ---------- KPIs ---------- */
function renderKPIs(){
  const box=$("#kpis");box.innerHTML="";
  const k=DATA.kpis;
  const items=[
    [k.people,"People"],[k.students,"Students"],[k.postdocs,"Postdocs"],[k.faculty,"Faculty / staff"],
    [k.themes,"Themes"],[k.sites,"Sites"],
    [k.quals,"Quals passed"],[k.comps,"Comps passed"],
    [k.publications,"Publications"],[k.presentations,"Presentations"],
    [k.awards,"Awards"],[k.grants,"Active grants"],
    [k.grant_total?("$"+Math.round(k.grant_total).toLocaleString()):0,"Funding on the books"],
  ].filter(([n])=>n);
  items.forEach(([n,l])=>{const d=el("div","kpi");d.append(el("div","n",n),el("div","l",l));box.append(d);});
}

/* ---------- timelines ---------- */
function timelineSVG(rows,opts){
  const W=1080,rowH=34,padL=170,padR=90,padT=34,padB=8;
  const H=padT+rows.length*rowH+padB;
  const x0=opts.min,x1=opts.max,span=(x1-x0)||1;
  const sx=v=>padL+((v-x0)/span)*(W-padL-padR);
  let s=`<svg viewBox="0 0 ${W} ${H}" role="img">`;
  s+=`<g class="axis">`;
  for(let y=Math.ceil(x0);y<=x1;y++){const X=sx(y);
    s+=`<line x1="${X}" y1="${padT-6}" x2="${X}" y2="${H-padB}" stroke="var(--rule)" stroke-width="1"/>`;
    s+=`<text x="${X}" y="${padT-12}" text-anchor="middle">${y}</text>`;}
  s+=`</g>`;
  if(opts.band){rows.forEach((r,i)=>{if(r.bandFrom==null)return;
    const y=padT+i*rowH+6,h=rowH-12,bx=sx(r.bandFrom),bw=sx(r.bandTo)-sx(r.bandFrom);
    s+=`<rect x="${bx}" y="${y}" width="${bw}" height="${h}" fill="var(--band)"/>`;});}
  const px=sx(NOW);
  s+=`<line x1="${px}" y1="${padT-6}" x2="${px}" y2="${H-padB}" stroke="var(--present)" stroke-width="1.4" stroke-dasharray="3,3"/>`;
  s+=`<text x="${px}" y="${H-padB}" text-anchor="middle" class="barcap" fill="var(--present)" dy="-1">now</text>`;
  rows.forEach((r,i)=>{const cy=padT+i*rowH+rowH/2,bx0=sx(r.from),bx1=sx(r.to);
    s+=`<text x="${padL-12}" y="${cy}" text-anchor="end" dominant-baseline="middle" class="rowlabel">${esc(r.label)}</text>`;
    s+=`<line x1="${padL}" y1="${cy}" x2="${W-padR}" y2="${cy}" stroke="var(--rule)" stroke-width="1" opacity=".5"/>`;
    s+=`<line x1="${bx0}" y1="${cy}" x2="${bx1}" y2="${cy}" stroke="${themeColor(r.key)}" stroke-width="6" stroke-linecap="round"/>`;
    s+=`<circle cx="${bx0}" cy="${cy}" r="3.4" fill="${themeColor(r.key)}"/>`;
    s+=`<text x="${bx1+8}" y="${cy}" dominant-baseline="middle" class="barcap">${esc(r.cap)}</text>`;});
  s+=`</svg>`;return s;
}
function renderCohort(){
  const rows=visiblePeople().filter(p=>isStudent(p)&&p.start_year!=null).map(p=>({
    label:`${p.name}${p.program_year?"  ·  "+p.program_year:""}`,key:p.key,from:p.start_year,to:NOW,
    bandFrom:p.start_year+5,bandTo:p.start_year+6,cap:`${(NOW-p.start_year).toFixed(1)} yr`}));
  const sec=$("#sec-cohort");
  if(!rows.length){sec.style.display="none";return;}sec.style.display="";
  const maxEnd=Math.max(...rows.map(r=>r.bandTo),NOW)+0.4;
  const minStart=Math.min(...rows.map(r=>r.from))-0.4;
  $("#cohortChart").innerHTML=timelineSVG(rows,{min:minStart,max:maxEnd,band:true});
}
function renderPostdoc(){
  const rows=visiblePeople().filter(p=>isPostdoc(p)&&p.start_year!=null).map(p=>({
    label:`${p.name}${p.effort?"  ·  "+p.effort:""}`,key:p.key,from:p.start_year,to:NOW,
    bandFrom:p.start_year+3,bandTo:Math.max(NOW,p.start_year+3.01),cap:`${(NOW-p.start_year).toFixed(1)} yr`}));
  const sec=$("#sec-postdoc");
  if(!rows.length){sec.style.display="none";return;}sec.style.display="";
  const minStart=Math.min(...rows.map(r=>r.from))-0.4;
  $("#postdocChart").innerHTML=timelineSVG(rows,{min:minStart,max:NOW+0.6,band:true});
}

/* ---------- people table ---------- */
let sortKey="name",sortDir=1;
function renderPeople(){
  const rows=visiblePeople();
  const cols=[["name","Name"],["rank","Rank"],["theme","Theme"],["program_year","Year"],
              ["site","Site"],["funding","Funding"]];
  rows.sort((a,b)=>String(a[sortKey]||"").localeCompare(String(b[sortKey]||""),undefined,{numeric:true})*sortDir);
  const t=el("table"),thead=el("thead"),trh=el("tr");
  cols.forEach(([k,l])=>{const th=el("th",null,l+(sortKey===k?(sortDir>0?" ▲":" ▼"):""));
    th.onclick=()=>{if(sortKey===k)sortDir*=-1;else{sortKey=k;sortDir=1;}renderPeople();};trh.append(th);});
  thead.append(trh);t.append(thead);
  const tb=el("tbody");
  rows.forEach(p=>{const tr=el("tr");
    tr.append(el("td",null,`<span class="theme-dot" style="background:${themeColor(p.key)}"></span>${esc(p.name)}`),
      el("td","muted",esc(p.rank)),el("td",null,esc(p.theme)),el("td",null,esc(p.program_year||"—")),
      el("td",null,esc(p.site||"—")),el("td","muted",esc(p.funding||"—")));
    tb.append(tr);});
  t.append(tb);$("#peopleTable").innerHTML="";$("#peopleTable").append(t);
  const lg=$("#legend");lg.innerHTML="";
  DATA.themes.forEach(t=>lg.append(el("span",null,`<span class="theme-dot" style="background:${t.color}"></span>${esc(t.label)}`)));
}

/* ---------- milestones ---------- */
function pill(v){
  const s=(v||"").toLowerCase();
  if(!s)return `<span class="pill na">—</span>`;
  if(s==="passed"||s==="yes"||s==="formed")return `<span class="pill ok">✓ ${esc(v)}</span>`;
  if(s==="n/a"||s==="na"||s==="no")return `<span class="pill na">${esc(v)}</span>`;
  return `<span class="pill pending">${esc(v)}</span>`;
}
function renderMilestones(){
  const rows=visiblePeople().filter(p=>isStudent(p));
  const sec=$("#sec-milestones");
  if(!rows.length){sec.style.display="none";return;}sec.style.display="";
  const t=el("table","mtable");
  t.innerHTML=`<thead><tr><th>Student</th><th>Qual</th><th>Comp</th><th>Committee</th><th>Author</th><th>Defense</th><th>In program</th></tr></thead>`;
  const tb=el("tbody");
  rows.forEach(p=>{const m=p.milestones||{};const g=k=>(m[k]||{}).status;
    const tr=el("tr");
    tr.innerHTML=`<td><span class="theme-dot" style="background:${themeColor(p.key)}"></span>${esc(p.name)} <span class="muted">${esc(p.program_year||"")}</span></td>`
      +`<td>${pill(g("qual"))}</td><td>${pill(g("comp"))}</td><td>${pill(g("committee"))}</td>`
      +`<td>${pill(g("author"))}</td><td>${pill(g("defense"))}</td>`
      +`<td class="muted">${p.start_year!=null?(NOW-p.start_year).toFixed(1)+" yr":"—"}</td>`;
    tb.append(tr);});
  t.append(tb);$("#milestoneTable").innerHTML="";$("#milestoneTable").append(t);
}

/* ---------- matrix ---------- */
function renderMatrix(){
  const rows=visiblePeople().filter(p=>(DATA.membership[p.name]));
  const projs=DATA.projects.filter(pr=>match(pr.key));
  const sec=$("#sec-matrix");
  if(!rows.length||!projs.length){sec.style.display="none";return;}sec.style.display="";
  const t=el("table"),thead=el("thead"),trh=el("tr");
  trh.append(el("th",null,"Person"));
  projs.forEach(pr=>{const th=el("th","proj",esc(pr.name));th.style.color=themeColor(pr.key);trh.append(th);});
  thead.append(trh);t.append(thead);
  const tb=el("tbody");
  rows.forEach(p=>{const tr=el("tr");
    tr.append(el("td","name",`<span class="theme-dot" style="background:${themeColor(p.key)}"></span>${esc(p.name)}`));
    projs.forEach(pr=>{const m=(DATA.membership[p.name]||{})[pr.name];const td=el("td");
      if(m)td.innerHTML=`<span class="cell ${m==='lead'?'lead':''}" style="background:${themeColor(pr.key)}" title="${esc(p.name)} — ${esc(pr.name)} (${m})"></span>`;
      tr.append(td);});
    tb.append(tr);});
  t.append(tb);$("#matrix").innerHTML="";$("#matrix").append(t);
}

/* ---------- funding board ---------- */
function renderFunding(){
  const board=$("#fundingBoard");board.innerHTML="";
  const vis=DATA.applications.filter(a=>match(a.key));
  const sec=$("#sec-funding");
  if(!vis.length){sec.style.display="none";return;}sec.style.display="";
  const order=[["Awarded","awarded"],["Under review","under-review"],["Applied","applied"],
               ["Planned","planned"],["Declined","declined"],["Not applying","planned"]];
  const seen={};vis.forEach(a=>{const s=(a.status||"Other").trim();(seen[s]=seen[s]||[]).push(a);});
  const cols=order.filter(([label])=>seen[label]);
  Object.keys(seen).forEach(s=>{if(!order.some(([l])=>l===s))cols.push([s,"planned"]);});
  cols.forEach(([label,cls])=>{
    const items=seen[label]||[];
    const col=el("div","col "+cls);
    col.append(el("h3",null,`<span>${esc(label)}</span><span>${items.length}</span>`));
    const box=el("div","items");
    items.forEach(a=>{const c=el("div","fcard");
      const amt=a.awarded||a.requested;
      c.append(el("div","who",esc(a.who)),
        el("div","what",esc(a.name)+(a.type?` · ${esc(a.type)}`:"")+(amt?` · ${esc(amt)}`:"")+(a.date?` · ${esc(a.date)}`:"")));
      box.append(c);});
    col.append(box);board.append(col);
  });
}

/* ---------- publications ---------- */
function renderPubs(){
  const rows=DATA.publications.filter(p=>match(p.key));
  const sec=$("#sec-pubs");
  if(!rows.length){sec.style.display="none";return;}sec.style.display="";
  const t=el("table");
  t.innerHTML=`<thead><tr><th>Author</th><th>Title</th><th>Role</th><th>Journal / collab.</th><th>Status</th><th>Date</th></tr></thead>`;
  const tb=el("tbody");
  rows.forEach(p=>{const tr=el("tr");
    const title=p.doi?`<a href="${esc(p.doi)}">${esc(p.title)}</a>`:esc(p.title);
    tr.innerHTML=`<td><span class="theme-dot" style="background:${themeColor(p.key)}"></span>${esc(p.who)}</td>`
      +`<td>${title}</td><td class="muted">${esc(p.role)}</td><td class="muted">${esc(p.journal)}</td>`
      +`<td>${pill(p.status)}</td><td class="muted">${esc(p.date)}</td>`;
    tb.append(tr);});
  t.append(tb);$("#pubsTable").innerHTML="";$("#pubsTable").append(t);
}

/* ---------- talks & awards ---------- */
function renderTalks(){
  const talks=DATA.presentations.filter(t=>match(t.key))
    .map(t=>({who:t.who,yr:t.date||"",label:`<strong>${esc(t.who)}</strong> — ${esc(t.type)}${t.venue?", "+esc(t.venue):""}${t.title?` · <em>${esc(t.title)}</em>`:""}`,
              prize:/yes|invited/i.test(t.invited||"")?"invited":""}));
  const awards=DATA.awards.filter(a=>match(a.key))
    .map(a=>({who:a.who,yr:a.date||"",label:`<strong>${esc(a.who)}</strong> — ${esc(a.award)}${a.org?", "+esc(a.org):""}`,
              prize:a.amount?esc(a.amount):"award"}));
  const rows=talks.concat(awards).sort((a,b)=>String(b.yr).localeCompare(String(a.yr)));
  const sec=$("#sec-talks");
  if(!rows.length){sec.style.display="none";return;}sec.style.display="";
  const box=$("#talks");box.innerHTML="";
  rows.forEach(r=>{const row=el("div","talk");
    row.append(el("div","yr",esc(r.yr)),el("div",null,r.label),el("div","prize",r.prize?"★ "+r.prize:""));
    box.append(row);});
}

/* ---------- grants & funding ---------- */
function renderGrants(){
  const rows=DATA.grants.filter(g=>match(g.key));
  const sec=$("#sec-grants");
  if(!rows.length){sec.style.display="none";return;}sec.style.display="";
  const rank={active:0,awarded:1,pending:2,submitted:3,planned:4,completed:5,declined:6};
  rows.sort((a,b)=>(rank[(a.status||"").toLowerCase()]??9)-(rank[(b.status||"").toLowerCase()]??9));
  const t=el("table");
  t.innerHTML=`<thead><tr><th>Holder</th><th>Grant / award</th><th>Agency</th><th>Role</th><th>Amount</th><th>Period</th><th>Status</th></tr></thead>`;
  const tb=el("tbody");
  rows.forEach(g=>{const tr=el("tr");
    const period=[g.start,g.end].filter(Boolean).join(" – ");
    tr.innerHTML=`<td><span class="theme-dot" style="background:${themeColor(g.key)}"></span>${esc(g.who)}</td>`
      +`<td>${esc(g.title)}</td><td class="muted">${esc(g.agency)}</td><td class="muted">${esc(g.role)}</td>`
      +`<td>${esc(g.amount)}</td><td class="muted">${esc(period)}</td><td>${pill(g.status)}</td>`;
    tb.append(tr);});
  t.append(tb);$("#grantsTable").innerHTML="";$("#grantsTable").append(t);
}

/* ---------- conferences & schools ---------- */
function renderEvents(){
  const rows=DATA.events.filter(e=>match(e.key));
  const sec=$("#sec-events");
  if(!rows.length){sec.style.display="none";return;}sec.style.display="";
  rows.sort((a,b)=>String(b.start||"").localeCompare(String(a.start||"")));
  const t=el("table");
  t.innerHTML=`<thead><tr><th>Person</th><th>Type</th><th>Event</th><th>Location</th><th>Date</th><th>Role</th><th>Presented?</th></tr></thead>`;
  const tb=el("tbody");
  rows.forEach(e=>{const tr=el("tr");
    const dstr=e.end&&e.end!==e.start?`${esc(e.start)} – ${esc(e.end)}`:esc(e.start);
    tr.innerHTML=`<td><span class="theme-dot" style="background:${themeColor(e.key)}"></span>${esc(e.who)}</td>`
      +`<td>${esc(e.type)}</td><td>${esc(e.event)}</td><td class="muted">${esc(e.location)}</td>`
      +`<td class="muted">${dstr}</td><td class="muted">${esc(e.role)}</td>`
      +`<td>${/yes/i.test(e.presented||"")?'<span class="pill ok">✓ yes</span>':'<span class="pill na">no</span>'}</td>`;
    tb.append(tr);});
  t.append(tb);$("#eventsTable").innerHTML="";$("#eventsTable").append(t);
}

/* ---------- generic record table ---------- */
function renderRecordTable(secId,tableId,rows,cols){
  const sec=$(secId);
  if(!rows.length){sec.style.display="none";return;}sec.style.display="";
  const t=el("table");
  t.innerHTML=`<thead><tr>${cols.map(c=>`<th>${c[1]}</th>`).join("")}</tr></thead>`;
  const tb=el("tbody");
  rows.forEach(r=>{const tr=el("tr");
    tr.innerHTML=cols.map((c,i)=>{
      if(i===0)return `<td><span class="theme-dot" style="background:${themeColor(r.key)}"></span>${esc(r[c[0]])}</td>`;
      return `<td class="${c[2]||''}">${esc(r[c[0]]||"")}</td>`;
    }).join("");
    tb.append(tr);});
  t.append(tb);$(tableId).innerHTML="";$(tableId).append(t);
}
function renderRoles(){
  renderRecordTable("#sec-roles","#rolesTable",DATA.roles.filter(r=>match(r.key)),
    [["who","Person"],["role","Role"],["category","Category","muted"],["org","Organization","muted"],
     ["start","Start","muted"],["end","End","muted"]]);
}
function renderOutreach(){
  renderRecordTable("#sec-outreach","#outreachTable",DATA.outreach.filter(r=>match(r.key)),
    [["who","Person"],["activity","Activity"],["type","Type","muted"],["audience","Audience / venue","muted"],
     ["date","Date","muted"],["role","Role","muted"]]);
}

/* ---------- render all ---------- */
function renderAll(){
  renderFilters();renderKPIs();renderCohort();renderPostdoc();renderPeople();
  renderMilestones();renderMatrix();renderGrants();renderFunding();renderPubs();
  renderTalks();renderEvents();renderRoles();renderOutreach();
}

$("#themeBtn").onclick=()=>{
  const cur=document.documentElement.getAttribute("data-theme")==="dark"?"light":"dark";
  document.documentElement.setAttribute("data-theme",cur);
};
if(window.matchMedia&&window.matchMedia("(prefers-color-scheme: dark)").matches)
  document.documentElement.setAttribute("data-theme","dark");

renderAll();
</script>
</body>
</html>
"""


if __name__ == "__main__":
    raise SystemExit(main())
