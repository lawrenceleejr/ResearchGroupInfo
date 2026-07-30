#!/usr/bin/env python3
"""Build/extend the Research Record template.

Idempotently updates ``template/FirstnameLastname_Info_Template.xlsx``:

  * adds "Faculty" to the Overview → Rank dropdown (so PIs/faculty can keep a
    record too);
  * adds a **Conferences & Schools** tab (attendance log — conferences,
    workshops, summer/winter schools, whether or not you presented);
  * adds a **Grants & Funding** tab (the group's funded grant portfolio,
    distinct from the per-person Applications tab);
  * orders the tabs sensibly.

New tabs are cloned from an existing styled sheet so they inherit the template
look (fonts, header band, example row, column widths), then relabeled and given
their own dropdowns.

Run:  python build_template.py
"""
import copy
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

PATH = "template/FirstnameLastname_Info_Template.xlsx"
DATA_ROWS = 40  # how many data rows the dropdowns cover

TITLE_ROW = 2
DESC_ROW = 5
HEADER_ROW = 6
EXAMPLE_ROW = 7


def dv(formula, cells):
    d = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
    d.add(cells)
    return d


def set_row(ws, row, values):
    for i, v in enumerate(values, start=1):
        ws.cell(row=row, column=i, value=v)


def clone_sheet(wb, src, new_name):
    if new_name in wb.sheetnames:
        del wb[new_name]
    ws = wb.copy_worksheet(wb[src])
    ws.title = new_name
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = "4E7C8B"
    return ws


def clear_data_validations(ws):
    # copy_worksheet doesn't carry DVs, but be safe on re-runs
    ws.data_validations.dataValidation = []


def build_conferences(wb):
    ws = clone_sheet(wb, "Presentations", "Conferences & Schools")
    clear_data_validations(ws)
    ws["A2"] = "CONFERENCES & SCHOOLS"
    ws["A5"] = ("Conferences, workshops, and summer/winter schools you attended "
                "— log every event, whether or not you presented.")
    set_row(ws, HEADER_ROW,
            ["#", "Type", "Event / meeting", "Location", "Start", "End",
             "Your role", "Presented?", "Notes"])
    set_row(ws, EXAMPLE_ROW,
            ["ex", "Summer school", "CERN–JINR European School", "Split, Croatia",
             "2025-09", "2025-09", "Student", "No", "poster session"])
    ws.add_data_validation(dv(
        "Conference,Workshop,Summer school,Winter school,School,"
        "Collaboration meeting,Seminar series,Other", f"B{EXAMPLE_ROW}:B{HEADER_ROW+DATA_ROWS}"))
    ws.add_data_validation(dv(
        "Attendee,Contributed talk,Invited talk,Poster,Lecturer,Organizer,"
        "Session chair,Other", f"G{EXAMPLE_ROW}:G{HEADER_ROW+DATA_ROWS}"))
    ws.add_data_validation(dv("Yes,No", f"H{EXAMPLE_ROW}:H{HEADER_ROW+DATA_ROWS}"))
    return ws


def build_grants(wb):
    ws = clone_sheet(wb, "Presentations", "Grants & Funding")
    clear_data_validations(ws)
    ws["A2"] = "GRANTS & FUNDING"
    ws["A5"] = ("Funded grants and awards supporting the group — the funding "
                "portfolio. (Individual fellowship/travel applications go on the "
                "Applications tab.)")
    set_row(ws, HEADER_ROW,
            ["#", "Grant / award title", "Agency / sponsor", "Your role",
             "Amount", "Start", "End", "Status", "Grant no. / notes"])
    set_row(ws, EXAMPLE_ROW,
            ["ex", "Muon collider R&D", "DOE Office of Science", "PI",
             "$750,000", "2024-08", "2027-07", "Active", "DE-SC00xxxxx"])
    ws.add_data_validation(dv(
        "PI,Co-PI,Senior personnel,Key personnel,Consultant,Other",
        f"D{EXAMPLE_ROW}:D{HEADER_ROW+DATA_ROWS}"))
    ws.add_data_validation(dv(
        "Active,Awarded,Pending,Submitted,Completed,Declined,Planned",
        f"H{EXAMPLE_ROW}:H{HEADER_ROW+DATA_ROWS}"))
    return ws


def update_rank_dropdown(wb):
    ws = wb["Overview"]
    for d in list(ws.data_validations.dataValidation):
        if d.formula1 and "Grad student" in d.formula1:
            d.formula1 = '"Undergrad,Grad student,Postdoc,Faculty,Research staff"'
    # refresh the year/graduation hints so faculty know to leave them N/A
    for r in range(1, ws.max_row + 1):
        a = ws[f"A{r}"].value
        if a and str(a).strip() == "Current program year":
            ws[f"E{r}"] = "e.g. Y3   (postdocs/faculty: N/A)"
        if a and str(a).strip() == "Expected graduation term":
            ws[f"E{r}"] = "target term   (postdocs/faculty: N/A)"


def reorder(wb):
    desired = ["Overview", "Committee", "Roles", "Presentations",
               "Conferences & Schools", "Outreach & education", "Publications",
               "Applications", "Grants & Funding", "Awards", "Notes"]
    order = [wb[n] for n in desired if n in wb.sheetnames]
    order += [ws for ws in wb.worksheets if ws not in order]
    wb._sheets = order


def main():
    wb = openpyxl.load_workbook(PATH)
    update_rank_dropdown(wb)
    build_conferences(wb)
    build_grants(wb)
    reorder(wb)
    wb.save(PATH)
    print("Updated", PATH)
    print("Tabs:", wb.sheetnames)


if __name__ == "__main__":
    main()
