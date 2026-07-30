#!/usr/bin/env python3
"""Generate a few filled-in sample Research Record workbooks under sample_data/.

These are illustrative only — they exercise every tab of the template so the
generated dashboard has something to show. Run once (or whenever the template
changes) to regenerate the demo inputs:

    python make_sample_data.py
"""
import os
import shutil
import openpyxl

TEMPLATE = "template/FirstnameLastname_Info_Template.xlsx"
OUT = "sample_data"

# label -> value for the Overview sheet (identity + research block)
PEOPLE = [
    {
        "file": "LawrenceLee_Info.xlsx",
        "overview": {
            "Rank": "Faculty", "Full name": "Lawrence Lee",
            "Research theme / area": "Collider (CMS)",
            "Program start (year / term)": "2019 Fall", "Started with the group": "2019 Fall",
            "Current program year": "N/A", "Expected graduation term": "N/A",
            "Appointment / effort": "full-time", "Primary site / base": "UTK",
            "Primary funding source": "NSF / DOE",
            "Primary project(s)": "Multijet BSM search; Muon cooling / collider",
            "Current focus / status": "Group PI — collider + accelerator programs",
        },
        "milestones": {},
        "grants": [
            ["1", "CAREER: Searches for new physics with jets", "NSF", "PI", "$700,000", "2021-08", "2026-07", "Active", "PHY-21xxxxx"],
            ["2", "Muon collider R&D", "DOE Office of Science", "Co-PI", "$1,200,000", "2024-08", "2027-07", "Active", "DE-SC00xxxxx"],
            ["3", "US-CMS operations program", "NSF", "Senior personnel", "$300,000", "2023-01", "2028-12", "Active", ""],
            ["4", "Detector R&D instrumentation", "DOE", "PI", "$450,000", "2026-01", "", "Pending", "under review"],
        ],
        "events": [
            ["1", "Conference", "APS April Meeting", "Sacramento, CA", "2025-04", "2025-04", "Session chair", "Yes", ""],
            ["2", "Collaboration meeting", "CMS Week", "CERN", "2026-03", "2026-03", "Attendee", "No", ""],
        ],
    },
    {
        "file": "JohnnyAppleseed_Info.xlsx",
        "overview": {
            "Rank": "Grad student", "Full name": "Johnny Appleseed",
            "Research theme / area": "Collider (CMS)",
            "Program start (year / term)": "2022 Fall", "Started with the group": "2022 Fall",
            "Current program year": "Y5", "Expected graduation term": "2027 Spring",
            "Appointment / effort": "full-time", "Primary site / base": "CERN",
            "Primary funding source": "NSF",
            "Working thesis title / topic": "Multijet resonance search",
            "Primary project(s)": "Multijet BSM search (lead); Outer tracker (lead)",
            "Current focus / status": "Finalizing multijet analysis for publication",
        },
        "milestones": {"Qualifying exam": ("Passed", "2023-05", "UTK", ""),
                       "Comprehensive exam": ("Passed", "2024-04", "UTK", ""),
                       "Thesis committee formed": ("Yes", "2024-05", "", ""),
                       "Author qualification (if applicable)": ("Passed", "2023-09", "CMS", ""),
                       "Thesis defense": ("Scheduled", "2027-04", "UTK", "")},
        "presentations": [
            ["1", "Contributed talk", "APS April Meeting", "Sacramento, CA", "2025-04", "Multijet search status", "No", 800, ""],
        ],
        "publications": [
            ["1", "Search for multijet resonances in pp collisions", "Analysis contact", "JHEP / CMS", "Under review", "2026-02", "10.xxxx/multijet"],
        ],
        "applications": [
            ["1", "Breakthrough Prize Fellowship", "Fellowship", "Applied", "", "", "2026-03", ""],
        ],
        "roles": [
            ["1", "Analysis contact", "Collaboration role", "CMS EXO", "2024-06", "ongoing", ""],
        ],
        "events": [
            ["1", "Conference", "APS April Meeting", "Sacramento, CA", "2025-04", "2025-04", "Contributed talk", "Yes", ""],
            ["2", "Summer school", "CMSDAS", "FNAL", "2023-01", "2023-01", "Attendee", "No", "data analysis school"],
        ],
    },
    {
        "file": "AdamBrown_Info.xlsx",
        "overview": {
            "Rank": "Grad student", "Full name": "Adam Brown",
            "Research theme / area": "Collider (CMS)",
            "Program start (year / term)": "2022 Fall", "Started with the group": "2024 Fall",
            "Current program year": "Y5", "Expected graduation term": "2027 Fall",
            "Appointment / effort": "full-time", "Primary site / base": "FNAL",
            "Primary funding source": "NSF CAREER",
            "Primary project(s)": "Outer tracker; HSCP search",
            "Current focus / status": "HSCP search + tracker commissioning",
        },
        "milestones": {"Qualifying exam": ("Passed", "2023-05", "UTK", ""),
                       "Comprehensive exam": ("Passed", "2024-11", "UTK", ""),
                       "Thesis committee formed": ("Yes", "2024-12", "", ""),
                       "Author qualification (if applicable)": ("Passed", "2024-01", "CMS", ""),
                       "Thesis defense": ("Not yet", "", "", "")},
        "applications": [
            ["1", "DOE SCGSR", "Fellowship", "Applied", "$50,000", "", "2026-05", "group-wide cycle"],
            ["2", "LPC Graduate Scholar", "Program", "Planned", "", "", "2026-09", "advisor-suggested"],
        ],
    },
    {
        "file": "CarolineDavis_Info.xlsx",
        "overview": {
            "Rank": "Grad student", "Full name": "Caroline Davis",
            "Research theme / area": "Accelerator",
            "Program start (year / term)": "2024 Fall", "Started with the group": "2024 Fall",
            "Current program year": "Y3", "Expected graduation term": "2029 Spring",
            "Appointment / effort": "full-time", "Primary site / base": "FNAL",
            "Primary funding source": "BNL / Cottrell Award",
            "Primary project(s)": "Muon cooling / collider (lead)",
            "Current focus / status": "Muon cooling demonstrator simulations",
        },
        "milestones": {"Qualifying exam": ("Passed", "2025-05", "UTK", ""),
                       "Comprehensive exam": ("Scheduled", "2026-11", "UTK", ""),
                       "Thesis committee formed": ("Scheduled", "", "", "after comp"),
                       "Author qualification (if applicable)": ("N/A", "", "", ""),
                       "Thesis defense": ("N/A", "", "", "")},
        "presentations": [
            ["1", "Contributed talk", "USMCC 2025", "Chicago, IL", "2025-06", "Muon cooling channel design", "No", 600, ""],
            ["2", "Poster", "USMCC 2025", "Chicago, IL", "2025-06", "Cooling cell optimization", "No", 0, ""],
            ["3", "Contributed talk", "IMCC 2026", "CERN", "2026-05", "Updated cooling results", "No", 1200, ""],
            ["4", "Lecture", "US Particle Accelerator School 2025", "Knoxville, TN", "2025-01", "Attendee — accelerator physics", "No", 0, ""],
        ],
        "events": [
            ["1", "Summer school", "US Particle Accelerator School", "Knoxville, TN", "2025-01", "2025-01", "Attendee", "No", "2-week course"],
            ["2", "Conference", "USMCC", "Chicago, IL", "2025-06", "2025-06", "Contributed talk", "Yes", ""],
            ["3", "Conference", "IMCC", "CERN", "2026-05", "2026-05", "Contributed talk", "Yes", ""],
        ],
        "awards": [
            ["1", "Best Poster Prize", "USMCC", "2025-06", "$500", ""],
            ["2", "Best Poster Prize", "DPF", "2026-05", "", ""],
        ],
        "applications": [
            ["1", "DOE SCGSR", "Fellowship", "Planned", "$50,000", "", "2026-05", "group-wide cycle"],
            ["2", "URA VSP", "Program", "Planned", "", "", "2026-06", ""],
        ],
        "outreach": [
            ["1", "Physics Open House demo", "Public outreach", "~200 visitors, FNAL", "2025-10", "Lead demonstrator", "4", "muon table", ""],
        ],
    },
    {
        "file": "EmeryClark_Info.xlsx",
        "overview": {
            "Rank": "Postdoc", "Full name": "Emery Clark",
            "Research theme / area": "Collider (CMS)",
            "Program start (year / term)": "2022 Spring", "Started with the group": "2022 Spring",
            "Current program year": "N/A", "Expected graduation term": "N/A",
            "Appointment / effort": "full-time", "Primary site / base": "CERN",
            "Primary funding source": "NSF / DOE / LPC-URA",
            "Primary project(s)": "Outer tracker",
            "Current focus / status": "CMS operations + tracker upgrade",
        },
        "milestones": {},
        "roles": [
            ["1", "Working-group convener", "Leadership", "CMS EXO", "2025-01", "ongoing", "co-convener"],
        ],
    },
    {
        "file": "DaisyEvans_Info.xlsx",
        "overview": {
            "Rank": "Postdoc", "Full name": "Daisy Evans",
            "Research theme / area": "Accelerator",
            "Program start (year / term)": "2025 Fall", "Started with the group": "2025 Fall",
            "Current program year": "N/A", "Expected graduation term": "N/A",
            "Appointment / effort": "75% (ramping)", "Primary site / base": "ORNL",
            "Primary funding source": "advisor's grant",
            "Primary project(s)": "Neutrino sim (muon collider) (lead)",
            "Current focus / status": "Neutrino simulation kickoff",
        },
        "milestones": {},
    },
]

OVERVIEW_MILESTONE_LABELS = {
    "Qualifying exam", "Comprehensive exam", "Thesis committee formed",
    "Author qualification (if applicable)", "Thesis defense",
}


def find_row(ws, label, col="A"):
    for r in range(1, ws.max_row + 1):
        v = ws[f"{col}{r}"].value
        if v is not None and str(v).strip().lower() == label.strip().lower():
            return r
    return None


def hash_row(ws):
    for r in range(1, ws.max_row + 1):
        if str(ws[f"A{r}"].value).strip() == "#":
            return r
    return None


def fill(person):
    wb = openpyxl.load_workbook(TEMPLATE)
    ov = wb["Overview"]
    for label, value in person["overview"].items():
        r = find_row(ov, label)
        if r:
            ov[f"B{r}"] = value
    for label, tup in person.get("milestones", {}).items():
        r = find_row(ov, label)
        if r:
            status, date, loc, notes = (list(tup) + ["", "", "", ""])[:4]
            ov[f"B{r}"], ov[f"C{r}"], ov[f"D{r}"], ov[f"E{r}"] = status, date, loc, notes

    def fill_table(sheet_name, rows):
        ws = wb[sheet_name]
        h = hash_row(ws)
        if h is None:
            return
        r = h + 2  # skip the 'ex' example row directly under the header
        for row in rows:
            for j, val in enumerate(row):
                ws.cell(row=r, column=j + 1, value=val)
            r += 1

    fill_table("Presentations", person.get("presentations", []))
    fill_table("Conferences & Schools", person.get("events", []))
    fill_table("Publications", person.get("publications", []))
    fill_table("Applications", person.get("applications", []))
    fill_table("Grants & Funding", person.get("grants", []))
    fill_table("Awards", person.get("awards", []))
    fill_table("Roles", person.get("roles", []))
    fill_table("Outreach & education", person.get("outreach", []))

    wb.save(os.path.join(OUT, person["file"]))


def main():
    if os.path.isdir(OUT):
        shutil.rmtree(OUT)
    os.makedirs(OUT)
    for person in PEOPLE:
        fill(person)
        print("wrote", person["file"])


if __name__ == "__main__":
    main()
