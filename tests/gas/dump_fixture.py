#!/usr/bin/env python3
"""Build the fixture for the Apps Script (Code.gs) tests.

Writes into the output directory (default: tests/gas/.fixture):
  sheets.json    every sample workbook's cell values, per tab
  expected.json  what the Python edition produces from the same records:
                 dashboard KPIs/themes and the five summary tables

The Node harness (gas.test.js) loads Code.gs with stubbed Google APIs, runs it
over sheets.json, and asserts the results match expected.json — keeping the
two editions provably in lockstep.
"""
import csv
import glob
import json
import os
import pathlib
import subprocess
import sys
import tempfile

ROOT = pathlib.Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))


def dump_sheets():
    import openpyxl
    out = {}
    for path in sorted(glob.glob(str(ROOT / "sample_data" / "*.xlsx"))):
        wb = openpyxl.load_workbook(path, data_only=True)
        sheets = {}
        for ws in wb.worksheets:
            maxc = ws.max_column
            rows = []
            for row in ws.iter_rows(values_only=True):
                vals = [("" if v is None else (v if isinstance(v, (int, float)) else str(v)))
                        for v in row]
                vals += [""] * (maxc - len(vals))
                rows.append(vals)
            sheets[ws.title] = rows
        out[os.path.basename(path)] = sheets
    return out


def dump_expected():
    with tempfile.TemporaryDirectory() as tmp:
        subprocess.run([sys.executable, str(ROOT / "generate_dashboard.py"),
                        str(ROOT / "sample_data"), "-o", f"{tmp}/d.html",
                        "--no-datestamp"], check=True, capture_output=True)
        html = open(f"{tmp}/d.html", encoding="utf-8").read()
        payload = html.split("const DATA = ", 1)[1].split(";\n", 1)[0]
        data = json.loads(payload.replace("<\\/", "</"))

        subprocess.run([sys.executable, str(ROOT / "generate_summaries.py"),
                        str(ROOT / "sample_data"), "-o", tmp,
                        "--no-datestamp"], check=True, capture_output=True)
        tables = {}
        for name in ["presentations", "publications", "conferences", "schools", "grants"]:
            with open(f"{tmp}/{name}.csv", newline="", encoding="utf-8") as f:
                tables[name] = list(csv.reader(f))
    return {
        "kpis": data["kpis"],
        "themes": [t["label"] for t in data["themes"]],
        "people": sorted(p["name"] for p in data["people"]),
        "tables": tables,
    }


def main():
    out_dir = pathlib.Path(sys.argv[1] if len(sys.argv) > 1
                           else pathlib.Path(__file__).parent / ".fixture")
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "sheets.json").write_text(json.dumps(dump_sheets()), encoding="utf-8")
    (out_dir / "expected.json").write_text(json.dumps(dump_expected()), encoding="utf-8")
    print(f"Fixture written to {out_dir}")


if __name__ == "__main__":
    main()
