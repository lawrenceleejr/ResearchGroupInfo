"""End-to-end tests: run the real scripts on the committed sample records."""
import filecmp
import json
import re
import shutil
import subprocess
import sys

import openpyxl
import pytest

import generate_dashboard as gd

EXPECTED_KPIS = {
    "people": 6, "students": 3, "postdocs": 2, "faculty": 1,
    "themes": 2, "sites": 4, "quals": 3, "comps": 2,
    "publications": 1, "presentations": 5, "awards": 2, "awarded": 0,
    "grants": 3, "grant_total": 2200000,
}


def run(script, *args, cwd=None):
    return subprocess.run([sys.executable, str(script), *map(str, args)],
                          capture_output=True, text=True, cwd=cwd, check=True)


def extract_data(html_path):
    html = html_path.read_text(encoding="utf-8")
    assert "/*__DATA__*/null" not in html
    payload = html.split("const DATA = ", 1)[1].split(";\n", 1)[0]
    return json.loads(payload.replace("<\\/", "</"))


# ---------------------------------------------------------------- dashboard
def test_dashboard_end_to_end(root, sample_dir, tmp_path):
    run(root / "generate_dashboard.py", sample_dir, "-o", tmp_path / "d.html",
        "--no-datestamp", "--title", "T")
    data = extract_data(tmp_path / "d.html")
    assert data["kpis"] == EXPECTED_KPIS
    assert sorted(t["label"] for t in data["themes"]) == ["Accelerator", "Collider (CMS)"]
    names = {p["name"] for p in data["people"]}
    assert names == {"Adam Brown", "Caroline Davis", "Daisy Evans",
                     "Emery Clark", "Johnny Appleseed", "Lawrence Lee"}
    # faculty stays out of the student cohort
    lee = next(p for p in data["people"] if p["name"] == "Lawrence Lee")
    assert "faculty" in lee["rank"].lower()


def test_dashboard_datestamps_filename(root, sample_dir, tmp_path):
    run(root / "generate_dashboard.py", sample_dir, "-o", tmp_path / "d.html")
    produced = [p.name for p in tmp_path.glob("d_*.html")]
    assert len(produced) == 1
    assert re.fullmatch(r"d_\d{4}-\d{2}-\d{2}\.html", produced[0])


def test_dashboard_creates_output_dir(root, sample_dir, tmp_path):
    out = tmp_path / "new" / "nested" / "d.html"
    run(root / "generate_dashboard.py", sample_dir, "-o", out, "--no-datestamp")
    assert out.exists()


# ---------------------------------------------------------------- summaries
def test_summaries_match_committed_demo(root, sample_dir, tmp_path):
    """Regression: the five CSVs must be identical to the committed demo output."""
    run(root / "generate_summaries.py", sample_dir, "-o", tmp_path, "--no-datestamp")
    for name in ["presentations.csv", "publications.csv", "conferences.csv",
                 "schools.csv", "grants.csv"]:
        assert filecmp.cmp(tmp_path / name, root / "summaries" / name,
                           shallow=False), f"{name} drifted from committed demo"


def test_summaries_datestamped_names(root, sample_dir, tmp_path):
    run(root / "generate_summaries.py", sample_dir, "-o", tmp_path)
    stamped = sorted(p.name for p in tmp_path.glob("*.csv"))
    assert len(stamped) == 5
    assert all(re.fullmatch(r"\w+_\d{4}-\d{2}-\d{2}\.csv", n) for n in stamped)


# ---------------------------------------------------------------- exclusions
def test_template_in_folder_is_not_a_member(root, sample_dir, tmp_path):
    records = tmp_path / "records"
    shutil.copytree(sample_dir, records)
    shutil.copy(root / "template" / "FirstnameLastname_Info_Template.xlsx", records)
    run(root / "generate_dashboard.py", records, "-o", tmp_path / "d.html",
        "--no-datestamp")
    data = extract_data(tmp_path / "d.html")
    assert data["kpis"]["people"] == 6  # template did not become a phantom member


# ---------------------------------------------------------------- parsing
def test_parse_workbook_fields(sample_dir):
    p = gd.parse_workbook(str(sample_dir / "CarolineDavis_Info.xlsx"))
    assert p["name"] == "Caroline Davis"
    assert p["rank"] == "Grad student"
    assert p["theme"] == "Accelerator"
    assert p["milestones"]["comp"]["status"] == "Scheduled"
    assert p["milestones"]["qual"]["status"] == "Passed"
    assert len(p["presentations"]) == 4
    assert len(p["awards"]) == 2
    assert len(p["events"]) == 3
    assert p["projects"] == [{"name": "Muon cooling / collider", "lead": True}]
    assert p["start_year"] == pytest.approx(2024.70)


def test_parse_workbook_grants(sample_dir):
    p = gd.parse_workbook(str(sample_dir / "LawrenceLee_Info.xlsx"))
    assert len(p["grants"]) == 4
    active = [g for g in p["grants"] if g["status"] == "Active"]
    assert sum(gd.parse_money(g["amount"]) for g in active) == 2200000


# ---------------------------------------------------------------- template
def test_template_structure(root):
    wb = openpyxl.load_workbook(root / "template" / "FirstnameLastname_Info_Template.xlsx")
    assert wb.sheetnames == [
        "Overview", "Committee", "Roles", "Presentations", "Conferences & Schools",
        "Outreach & education", "Publications", "Applications", "Grants & Funding",
        "Awards", "Notes",
    ]
    ov = wb["Overview"]
    labels = {str(ov[f"A{r}"].value).strip() for r in range(1, ov.max_row + 1)
              if ov[f"A{r}"].value}
    for required in ["Rank", "Full name", "Research theme / area",
                     "Program start (year / term)", "Primary project(s)",
                     "Qualifying exam", "Comprehensive exam", "Thesis defense"]:
        assert required in labels, f"Overview label missing: {required}"
    rank_dv = [dv for dv in ov.data_validations.dataValidation
               if dv.formula1 and "Grad student" in dv.formula1]
    assert rank_dv and "Faculty" in rank_dv[0].formula1

    # every log tab has a '#' header row; Awards is free-text (no dropdowns)
    for tab in ["Presentations", "Conferences & Schools", "Grants & Funding",
                "Publications", "Applications", "Awards", "Roles"]:
        ws = wb[tab]
        hash_rows = [r for r in range(1, ws.max_row + 1)
                     if str(ws[f"A{r}"].value).strip() == "#"]
        assert hash_rows, f"{tab}: no '#' header row"
    for tab in ["Presentations", "Conferences & Schools", "Grants & Funding",
                "Publications", "Applications", "Roles"]:
        assert wb[tab].data_validations.dataValidation, f"{tab}: dropdowns missing"


def test_build_template_idempotent(root, tmp_path):
    (tmp_path / "template").mkdir()
    shutil.copy(root / "template" / "FirstnameLastname_Info_Template.xlsx",
                tmp_path / "template")
    run(root / "build_template.py", cwd=tmp_path)
    wb = openpyxl.load_workbook(tmp_path / "template" / "FirstnameLastname_Info_Template.xlsx")
    assert len(wb.sheetnames) == 11
    assert wb.sheetnames.count("Conferences & Schools") == 1
    assert wb.sheetnames.count("Grants & Funding") == 1


def test_make_sample_data(root, tmp_path):
    (tmp_path / "template").mkdir()
    shutil.copy(root / "template" / "FirstnameLastname_Info_Template.xlsx",
                tmp_path / "template")
    run(root / "make_sample_data.py", cwd=tmp_path)
    files = sorted(p.name for p in (tmp_path / "sample_data").glob("*.xlsx"))
    assert len(files) == 6
    p = gd.parse_workbook(str(tmp_path / "sample_data" / "AdamBrown_Info.xlsx"))
    assert p["name"] == "Adam Brown"
